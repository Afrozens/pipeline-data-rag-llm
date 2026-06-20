from typing import Dict, List, Optional, Tuple, Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

def _get_worksheet(file_path: str, sheet_name: Optional[str] = None) -> Tuple[Worksheet, str]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            ws = wb[sheet_name]
        else:
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
        return ws, sheet_name
    finally:
        wb.close()


def get_sheet_names(file_path: str) -> List[str]:
    wb = load_workbook(file_path, read_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def read_excel_headers(file_path: str, sheet_name: Optional[str] = None) -> List[str]:
    ws, _ = _get_worksheet(file_path, sheet_name)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return headers


def read_excel_sample(
    file_path: str,
    sheet_name: Optional[str] = None,
    n: int = 5,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    ws, _ = _get_worksheet(file_path, sheet_name)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= n:
            break
        rows.append(dict(zip(headers, row)))
    return headers, rows


def read_excel_data(
    file_path: str,
    sheet_name: Optional[str] = None,
) -> Tuple[List[str], List[Dict[str, Any]], int]:
    ws, _ = _get_worksheet(file_path, sheet_name)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
        total += 1
    return headers, rows, total
